# leads/management/commands/import_leads_xlsx.py
"""
One-off seed: import the candidate-bibliography spreadsheet into DatasetLead.

Idempotent — re-runs match on (title, volume_example) and update in place, so
running it twice does not create duplicates. These rows are the team's own
research, so they land as provenance=OWN, status=TRIAGE.

Usage (inside the web container):
    ./manage.py import_leads_xlsx /tmp/WHG_gazetteer_bibliography.xlsx
    ./manage.py import_leads_xlsx <path> --sheet "Candidate Bibliography" --dry-run
"""
from django.core.management.base import BaseCommand, CommandError

from leads.models import DatasetLead, LeadProvenance, LeadStatus, ScanStatus

DEFAULT_SHEET = 'Candidate Bibliography'

# Spreadsheet header label (normalised: lower-cased, collapsed whitespace) -> model field.
HEADER_MAP = {
    'title': 'title',
    'volume / district example': 'volume_example',
    'author / compiler': 'author_compiler',
    'publication year(s)': 'publication_years',
    'region covered': 'region_covered',
    'current country/area': 'current_area',
    'repository checked': 'repository',
    'source / access url': 'source_url',
    'pdf / scan status': '_scan_status_raw',   # free text -> mapped to enum below
    'suggested next action': 'next_action',
    'tags': '_tags_raw',                        # "a; b; c" -> list
}

MAX_LEN = {
    'title': 500, 'volume_example': 500, 'author_compiler': 500,
    'publication_years': 100, 'current_area': 255, 'repository': 255,
    'source_url': 1000,
}


def _norm(s):
    return ' '.join(str(s).strip().lower().split()) if s is not None else ''


def _map_scan_status(raw):
    """Heuristically map the free-text 'PDF / scan status' column to a ScanStatus."""
    t = (raw or '').lower()
    if not t.strip():
        return ScanStatus.UNKNOWN
    if any(k in t for k in ('pdf', 'download', 'scans available', 'scan available', 'full text')):
        return ScanStatus.DOWNLOADABLE
    if any(k in t for k in ('full view', 'fullview', 'catalog', 'catalogue', 'online edition', 'searchable')):
        return ScanStatus.FULLVIEW
    if any(k in t for k in ('not digit', 'no scan', 'no digit', 'undigit')):
        return ScanStatus.NONE
    return ScanStatus.UNKNOWN


def _split_tags(raw):
    if not raw:
        return []
    # tags are separated by ';' in the sheet; tolerate stray commas too
    parts = []
    for chunk in str(raw).replace('\n', ';').split(';'):
        parts.extend(p.strip() for p in chunk.split(',') if p.strip())
    # de-dupe, preserve order
    seen, out = set(), []
    for p in parts:
        if p.lower() not in seen:
            seen.add(p.lower())
            out.append(p)
    return out


class Command(BaseCommand):
    help = 'Seed DatasetLead rows from the candidate-bibliography .xlsx (idempotent).'

    def add_arguments(self, parser):
        parser.add_argument('xlsx_path', help='Path to WHG_gazetteer_bibliography.xlsx')
        parser.add_argument('--sheet', default=DEFAULT_SHEET,
                            help=f'Worksheet name (default: "{DEFAULT_SHEET}")')
        parser.add_argument('--dry-run', action='store_true',
                            help='Parse and report without writing to the database.')

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError as e:
            raise CommandError(f'openpyxl is required: {e}')

        path = opts['xlsx_path']
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'File not found: {path}')
        except Exception as e:  # noqa: BLE001
            raise CommandError(f'Could not open workbook: {e}')

        sheet = opts['sheet']
        if sheet not in wb.sheetnames:
            raise CommandError(f'Sheet "{sheet}" not found. Available: {wb.sheetnames}')
        ws = wb[sheet]

        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            raise CommandError(f'Sheet "{sheet}" is empty.')

        # Resolve column index -> model field via the normalised header map.
        col_field = {}
        for idx, label in enumerate(header):
            field = HEADER_MAP.get(_norm(label))
            if field:
                col_field[idx] = field
        if 'title' not in col_field.values():
            raise CommandError(f'No "title" column found in sheet "{sheet}". Header: {header}')

        dry = opts['dry_run']
        created = updated = skipped = 0

        for raw_row in rows:
            row = {}
            for idx, field in col_field.items():
                val = raw_row[idx] if idx < len(raw_row) else None
                row[field] = ('' if val is None else str(val).strip())

            title = row.get('title', '').strip()
            if not title:
                skipped += 1
                continue

            scan_status = _map_scan_status(row.pop('_scan_status_raw', ''))
            scan_raw = next((str(raw_row[i]).strip() for i, f in col_field.items()
                             if f == '_scan_status_raw' and raw_row[i]), '')
            tags = _split_tags(row.pop('_tags_raw', ''))

            defaults = {k: v for k, v in row.items() if not k.startswith('_') and k != 'volume_example'}
            # truncate to model max_length to be safe
            for k, limit in MAX_LEN.items():
                if k in defaults and defaults[k]:
                    defaults[k] = defaults[k][:limit]
            defaults.update({
                'scan_status': scan_status,
                'tags': tags,
                'provenance': LeadProvenance.OWN,
                'status': LeadStatus.TRIAGE,
            })
            # preserve the raw scan-status text (no dedicated field) in notes
            if scan_raw:
                defaults['notes'] = f'Scan status (from import): {scan_raw}'

            volume = row.get('volume_example', '')[:MAX_LEN['volume_example']]

            if dry:
                created += 1  # count as "would create/update"
                continue

            obj, was_created = DatasetLead.objects.update_or_create(
                title=title[:MAX_LEN['title']],
                volume_example=volume,
                defaults=defaults,
            )
            created += int(was_created)
            updated += int(not was_created)

        if dry:
            self.stdout.write(self.style.WARNING(
                f'[dry-run] {created} row(s) would be imported from "{sheet}"; {skipped} blank row(s) skipped.'))
        else:
            self.stdout.write(self.style.SUCCESS(
                f'Imported "{sheet}": {created} created, {updated} updated, {skipped} blank skipped. '
                f'Total leads now: {DatasetLead.objects.count()}.'))
